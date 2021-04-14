from openpyxl import load_workbook
from robot.api.deco import keyword


@keyword
def read_excel() -> list:
    wb = load_workbook('excel.xlsx')
    sheet = wb.active
    excel_data = []
    for row in range(2, 12):  # 10 rows in challenge

        data_dict = {
            'name_first': get_cell_value(sheet, row, 1),
            'name_last': get_cell_value(sheet, row, 2),
            'company': get_cell_value(sheet, row, 3),
            'role': get_cell_value(sheet, row, 4),
            'address': get_cell_value(sheet, row, 5),
            'email': get_cell_value(sheet, row, 6),
            'phone': get_cell_value(sheet, row, 7)
        }
        excel_data.append(data_dict)
    return excel_data


def get_cell_value(sheet, row, column) -> str:
    return str(sheet.cell(row=row, column=column).value)
